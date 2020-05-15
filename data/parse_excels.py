# - coding: utf-8 --
'''
@info create metro database from excels
adapted from Prof. Rui Henriques' code
'''
import openpyxl.utils

path = 'parsed_data/full_data/'
tablename = 'metro'
stationIDs = 'base_data/stops/stops.txt'
#dataFiles = ['metro/saidas_outubro_2018.xlsx','metro/entradas_outubro_2018.xlsx'] 
dataFiles = ['base_data/metro/entradas_outubro_2018.xlsx'] 
head = {
    'id':'serial primary key',
    'linha_entrada':'varchar(8)',
    'estacao_entrada':'varchar(4)',
    'linha_saida':'varchar(8)',
    'estacao_saida':'varchar(4)',
    'date':'timestamp',
    'count':'integer'}


def transform_data():

    #A: create temporary file
    temporary = path+'samplemetro_only_entries_station_names.csv'
    stations = get_station_ids()
    with open(temporary, 'w') as f:
        f.write(','.join(head.keys()) + "\n")
        
        #B: for each excel data file (entries and exits)
        for dataFile in dataFiles:
            print(dataFile)
            isentry = 'entradas' in dataFile
            wb = openpyxl.load_workbook(dataFile, read_only=True)
            
            #C: for each day (excel sheet) write records in temporary file
            for sheetname in wb.sheetnames:
                print('Sheet:'+sheetname)
                write_sheet_to_file(f, wb[sheetname], int(sheetname), isentry, stations)
            
    return [temporary]


def get_station_ids():
    stationdic = {}
    with open(stationIDs, 'r', encoding="utf8") as lines:
        next(lines)
        for line in lines:
            entries = line[:-1].split(',')
            #stationdic[entries[2].lower()]=entries[0]
            stationdic[entries[2].lower()]=entries[2].lower()
    return stationdic

def get_station_id(stations, name):
    station = name.replace(' 1','').replace(' 2','').replace('-',' ').lower()
    if station in stations: return stations[station]
    else: return None

def get_time(time, day):
    try: 
        if int(time[0:2])<4: day=day+1
    except: pass
    yearmonth = '2018-10-' if day<=31 else "2018-11-"
    if day>31: day=day%31
    daystr = '%d'%day
    if day<10: daystr = '0'+daystr
    if '-' in time: return yearmonth+daystr+' '+time[:-3]+':00'
    else: return None #yearmonth+daystr+' '+'11:11:11'

def write_sheet_to_file(writer, wsheet, day, isentry, stations):
    
    #A: init variables and fix table
    k = 1 if isentry else 0
    exit_line,entry_line,entry_station,exit_station,time = '.','.','.','.','.'
    limitcell = '%s%d' % (openpyxl.utils.get_column_letter(wsheet.max_column), wsheet.max_row)
    sheet = wsheet['G14':limitcell]
    
    #B: write table content
    for i in range(3+k,wsheet.max_row-14):
        if sheet[i][1-k].value!='': entry_line = sheet[i][1-k].value
        if sheet[i][2-k].value!='': entry_station = get_station_id(stations,sheet[i][2-k].value)
        if entry_station is None: continue
        time = get_time(sheet[i][3-k].value.replace(' ',''), day)
        if time is None: continue
        for j in range(4-k,65-k):
            if sheet[1+k][j].value!='': exit_line = sheet[1+k][j].value #[6:]
            exit_station = get_station_id(stations,sheet[2+k][j].value)
            if exit_station is None: continue
            if sheet[i][j].value is not None and sheet[i][j].value!='':
                line = ('%d,' % k)+entry_line+','+entry_station+','+exit_line+','+exit_station+','+time+(',%d\n' % sheet[i][j].value)
                writer.write(line.replace('Linha ','').replace('í','i').replace('ã','a'))


transform_data()